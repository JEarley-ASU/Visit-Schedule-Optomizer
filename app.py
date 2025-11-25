import streamlit as st
import pandas as pd
import numpy as np
import json
import s_optomize as sop
from io import BytesIO

# Page configuration
st.set_page_config(page_title="Meeting Scheduler", layout="wide")

# Initialize session state
if 'num_slots' not in st.session_state:
    st.session_state.num_slots = 6
if 'professor_data' not in st.session_state:
    st.session_state.professor_data = pd.DataFrame(columns=[f"Slot {i+1}" for i in range(st.session_state.num_slots)])
if 'visitor_data' not in st.session_state:
    st.session_state.visitor_data = pd.DataFrame(columns=[f"Preference {i+1}" for i in range(st.session_state.num_slots+3)])
if 'initialized' not in st.session_state:
    st.session_state.initialized = False
if 'name_memory' not in st.session_state:
    st.session_state.name_memory = []
if 'visitor_schedule' not in st.session_state:
    st.session_state.visitor_schedule = None
if 'professor_schedule' not in st.session_state:
    st.session_state.professor_schedule = None
if 'filename' not in st.session_state:
    st.session_state.filename = None

def initialize_schedule(new_num_slots):
    """Initialize or update the schedule with new number of slots."""
    old_num_slots = st.session_state.num_slots
    
    # If number of slots changed, clear old schedules as they're no longer valid
    if old_num_slots != new_num_slots:
        st.session_state.visitor_schedule = None
        st.session_state.professor_schedule = None
        if 'preference_analysis' in st.session_state:
            del st.session_state.preference_analysis
    
    # Update professor data columns
    new_prof_columns = [f"Slot {i + 1}" for i in range(new_num_slots)]
    if st.session_state.professor_data.empty:
        st.session_state.professor_data = pd.DataFrame(columns=new_prof_columns)
    else:
        if old_num_slots != new_num_slots:
            if new_num_slots > old_num_slots:
                for i in range(old_num_slots, new_num_slots):
                    st.session_state.professor_data[f"Slot {i + 1}"] = 1
            else:
                for i in range(new_num_slots, old_num_slots):
                    if f"Slot {i + 1}" in st.session_state.professor_data.columns:
                        st.session_state.professor_data = st.session_state.professor_data.drop(columns=[f"Slot {i + 1}"])
    
    # Update visitor data columns
    new_visitor_columns = [f"Preference {i + 1}" for i in range(new_num_slots + 3)]
    if st.session_state.visitor_data.empty:
        st.session_state.visitor_data = pd.DataFrame(columns=new_visitor_columns)
    else:
        if old_num_slots != new_num_slots:
            old_pref_cols = old_num_slots + 3
            new_pref_cols = new_num_slots + 3
            
            if new_pref_cols > old_pref_cols:
                for i in range(old_pref_cols, new_pref_cols):
                    st.session_state.visitor_data[f"Preference {i + 1}"] = " "
            else:
                for i in range(new_pref_cols, old_pref_cols):
                    if f"Preference {i + 1}" in st.session_state.visitor_data.columns:
                        st.session_state.visitor_data = st.session_state.visitor_data.drop(columns=[f"Preference {i + 1}"])
    
    st.session_state.num_slots = new_num_slots
    st.session_state.initialized = True
    st.rerun()

def add_professor(first_name, last_name, unavailable_slots):
    """Add a new professor to the data."""
    name = f"{last_name.upper()}, {first_name.upper()}"
    
    # Initialize availability
    availability = [1] * st.session_state.num_slots
    for slot in unavailable_slots:
        availability[slot] = 0
    
    # Update DataFrame
    if st.session_state.professor_data.empty:
        st.session_state.professor_data = pd.DataFrame(columns=[f"Slot {i + 1}" for i in range(st.session_state.num_slots)])
    
    st.session_state.professor_data.loc[name] = availability
    # Check if professor is completely unavailable and clear preferences
    clear_preferences_for_unavailable_professors()
    st.rerun()

def remove_professor(professor_name):
    """Remove a professor from the data."""
    if professor_name in st.session_state.professor_data.index:
        st.session_state.professor_data = st.session_state.professor_data.drop(index=professor_name)
        st.rerun()

def clear_preferences_for_unavailable_professors():
    """Clear visitor preferences for professors who are completely unavailable."""
    if st.session_state.professor_data.empty or st.session_state.visitor_data.empty:
        return
    
    # Find professors who are completely unavailable
    unavailable_professors = []
    for prof_name in st.session_state.professor_data.index:
        prof_row = st.session_state.professor_data.loc[prof_name]
        if all(prof_row[col] == 0 for col in prof_row.index):
            unavailable_professors.append(prof_name)
    
    # Clear preferences for unavailable professors
    if unavailable_professors:
        for visitor in st.session_state.visitor_data.index:
            for col in st.session_state.visitor_data.columns:
                pref_value = st.session_state.visitor_data.loc[visitor, col]
                if pref_value in unavailable_professors:
                    st.session_state.visitor_data.loc[visitor, col] = " "

def add_visitor(first_name, last_name):
    """Add a new visitor to the data."""
    if first_name and last_name:
        # Format as LAST, FIRST
        visitor_name = f"{last_name.upper()}, {first_name.upper()}"
        
        # Ensure the DataFrame has the correct columns
        expected_columns = [f"Preference {i+1}" for i in range(st.session_state.num_slots + 3)]
        
        # If columns don't match, recreate the DataFrame with correct columns
        if list(st.session_state.visitor_data.columns) != expected_columns:
            # Save existing data
            existing_data = st.session_state.visitor_data.copy()
            # Recreate with correct columns
            st.session_state.visitor_data = pd.DataFrame(columns=expected_columns)
            # Restore existing data (if any columns match)
            for idx in existing_data.index:
                row_data = {}
                for col in expected_columns:
                    if col in existing_data.columns:
                        row_data[col] = existing_data.loc[idx, col]
                    else:
                        row_data[col] = " "
                st.session_state.visitor_data.loc[idx] = row_data
        
        # Create default preferences as a dictionary matching column names
        default_preferences = {f"Preference {i+1}": " " for i in range(st.session_state.num_slots + 3)}
        st.session_state.visitor_data.loc[visitor_name] = default_preferences
        if visitor_name not in st.session_state.name_memory:
            st.session_state.name_memory.append(visitor_name)
        st.rerun()

def remove_visitor(visitor_name):
    """Remove a visitor from the data."""
    if visitor_name in st.session_state.visitor_data.index:
        st.session_state.visitor_data = st.session_state.visitor_data.drop(index=visitor_name)
        if visitor_name in st.session_state.name_memory:
            st.session_state.name_memory.remove(visitor_name)
        st.rerun()

def import_visitors_from_excel(uploaded_file):
    """Import visitor names from an Excel file."""
    try:
        # Read the Excel file
        df = pd.read_excel(uploaded_file)
        
        # Try to find first and last name columns (case-insensitive)
        first_name_col = None
        last_name_col = None
        name_col = None
        
        # Check for common column name variations
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in ['first name', 'firstname', 'first', 'fname', 'given name', 'givenname']:
                first_name_col = col
            elif col_lower in ['last name', 'lastname', 'last', 'lname', 'surname', 'family name', 'familyname']:
                last_name_col = col
            elif col_lower in ['name', 'full name', 'fullname', 'visitor', 'visitor name']:
                name_col = col
        
        imported_count = 0
        skipped_count = 0
        errors = []
        visitors_to_add = []
        
        # Ensure the DataFrame has the correct columns
        expected_columns = [f"Preference {i+1}" for i in range(st.session_state.num_slots + 3)]
        if list(st.session_state.visitor_data.columns) != expected_columns:
            existing_data = st.session_state.visitor_data.copy()
            st.session_state.visitor_data = pd.DataFrame(columns=expected_columns)
            for idx in existing_data.index:
                row_data = {}
                for col in expected_columns:
                    if col in existing_data.columns:
                        row_data[col] = existing_data.loc[idx, col]
                    else:
                        row_data[col] = " "
                st.session_state.visitor_data.loc[idx] = row_data
        
        # Case 1: Separate first and last name columns
        if first_name_col and last_name_col:
            for idx, row in df.iterrows():
                first_name = str(row[first_name_col]).strip() if pd.notna(row[first_name_col]) else ""
                last_name = str(row[last_name_col]).strip() if pd.notna(row[last_name_col]) else ""
                
                if first_name and last_name and first_name.lower() != 'nan' and last_name.lower() != 'nan':
                    visitor_name = f"{last_name.upper()}, {first_name.upper()}"
                    if visitor_name not in st.session_state.visitor_data.index:
                        visitors_to_add.append((first_name, last_name, visitor_name))
                        imported_count += 1
                    else:
                        skipped_count += 1
                else:
                    errors.append(f"Row {idx + 2}: Missing first or last name")
        
        # Case 2: Single name column (try to split by comma or space)
        elif name_col:
            for idx, row in df.iterrows():
                full_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                
                if full_name and full_name.lower() != 'nan':
                    # Try splitting by comma first (LAST, FIRST format)
                    if ',' in full_name:
                        parts = [p.strip() for p in full_name.split(',', 1)]
                        if len(parts) == 2:
                            last_name, first_name = parts[0], parts[1]
                        else:
                            # If only one part after comma, treat as last name
                            last_name, first_name = parts[0], ""
                    else:
                        # Split by space (assume FIRST LAST format)
                        parts = full_name.split()
                        if len(parts) >= 2:
                            first_name = parts[0]
                            last_name = ' '.join(parts[1:])
                        else:
                            first_name = parts[0] if parts else ""
                            last_name = ""
                    
                    if first_name and last_name:
                        visitor_name = f"{last_name.upper()}, {first_name.upper()}"
                        if visitor_name not in st.session_state.visitor_data.index:
                            visitors_to_add.append((first_name, last_name, visitor_name))
                            imported_count += 1
                        else:
                            skipped_count += 1
                    else:
                        errors.append(f"Row {idx + 2}: Could not parse name '{full_name}'")
                else:
                    errors.append(f"Row {idx + 2}: Empty name")
        
        # Case 3: No recognized columns - try first two columns
        else:
            if len(df.columns) >= 2:
                # Assume first column is first name, second is last name
                first_name_col = df.columns[0]
                last_name_col = df.columns[1]
                for idx, row in df.iterrows():
                    first_name = str(row[first_name_col]).strip() if pd.notna(row[first_name_col]) else ""
                    last_name = str(row[last_name_col]).strip() if pd.notna(row[last_name_col]) else ""
                    
                    if first_name and last_name and first_name.lower() != 'nan' and last_name.lower() != 'nan':
                        visitor_name = f"{last_name.upper()}, {first_name.upper()}"
                        if visitor_name not in st.session_state.visitor_data.index:
                            visitors_to_add.append((first_name, last_name, visitor_name))
                            imported_count += 1
                        else:
                            skipped_count += 1
                    else:
                        errors.append(f"Row {idx + 2}: Missing first or last name")
            else:
                return False, "Could not find name columns. Expected columns: 'First Name'/'Last Name' or 'Name'", None
        
        # Add all visitors in batch
        for first_name, last_name, visitor_name in visitors_to_add:
            default_preferences = {f"Preference {i+1}": " " for i in range(st.session_state.num_slots + 3)}
            st.session_state.visitor_data.loc[visitor_name] = default_preferences
            if visitor_name not in st.session_state.name_memory:
                st.session_state.name_memory.append(visitor_name)
        
        # Prepare result message
        result_msg = f"Successfully imported {imported_count} visitor(s)."
        if skipped_count > 0:
            result_msg += f" Skipped {skipped_count} duplicate(s)."
        if errors:
            result_msg += f" {len(errors)} error(s) encountered."
        
        return True, result_msg, errors if errors else None
        
    except Exception as e:
        return False, f"Error reading Excel file: {str(e)}", None

def save_state():
    """Save the current state to a JSON string."""
    # Convert DataFrames to dictionaries with index preserved
    prof_data_dict = st.session_state.professor_data.to_dict('index') if not st.session_state.professor_data.empty else {}
    visitor_data_dict = st.session_state.visitor_data.to_dict('index') if not st.session_state.visitor_data.empty else {}
    
    # Save schedules if they exist
    visitor_schedule_dict = None
    if st.session_state.visitor_schedule is not None and not st.session_state.visitor_schedule.empty:
        visitor_schedule_dict = st.session_state.visitor_schedule.to_dict('index')
    
    professor_schedule_dict = None
    if st.session_state.professor_schedule is not None and not st.session_state.professor_schedule.empty:
        professor_schedule_dict = st.session_state.professor_schedule.to_dict('index')
    
    preference_analysis_dict = None
    if 'preference_analysis' in st.session_state and st.session_state.preference_analysis is not None and not st.session_state.preference_analysis.empty:
        preference_analysis_dict = st.session_state.preference_analysis.to_dict('index')
    
    state = {
        'num_slots': st.session_state.num_slots,
        'professor_data': prof_data_dict,
        'visitor_data': visitor_data_dict,
        'name_memory': st.session_state.name_memory,
        'initialized': st.session_state.initialized,
        'max_students': st.session_state.get('max_students', 1),
        'visitor_schedule': visitor_schedule_dict,
        'professor_schedule': professor_schedule_dict,
        'preference_analysis': preference_analysis_dict,
        'filename': st.session_state.get('filename', None)
    }
    return json.dumps(state, indent=2, default=str)

def load_state(state_json):
    """Load state from a JSON string."""
    try:
        state = json.loads(state_json)
        st.session_state.num_slots = state['num_slots']
        
        # Load DataFrames with index preserved
        if state.get('professor_data'):
            st.session_state.professor_data = pd.DataFrame.from_dict(state['professor_data'], orient='index')
        else:
            st.session_state.professor_data = pd.DataFrame(columns=[f"Slot {i+1}" for i in range(st.session_state.num_slots)])
        
        if state.get('visitor_data'):
            st.session_state.visitor_data = pd.DataFrame.from_dict(state['visitor_data'], orient='index')
        else:
            st.session_state.visitor_data = pd.DataFrame(columns=[f"Preference {i+1}" for i in range(st.session_state.num_slots+3)])
        
        st.session_state.name_memory = state.get('name_memory', [])
        st.session_state.initialized = state.get('initialized', False)
        st.session_state.max_students = state.get('max_students', 1)
        
        # Load schedules if they exist
        if state.get('visitor_schedule') is not None:
            st.session_state.visitor_schedule = pd.DataFrame.from_dict(state['visitor_schedule'], orient='index')
        else:
            st.session_state.visitor_schedule = None
            
        if state.get('professor_schedule') is not None:
            st.session_state.professor_schedule = pd.DataFrame.from_dict(state['professor_schedule'], orient='index')
        else:
            st.session_state.professor_schedule = None
            
        if state.get('preference_analysis') is not None:
            st.session_state.preference_analysis = pd.DataFrame.from_dict(state['preference_analysis'], orient='index')
        else:
            if 'preference_analysis' in st.session_state:
                del st.session_state.preference_analysis
        
        st.session_state.filename = state.get('filename', None)
        
        st.rerun()
    except Exception as e:
        st.error(f"Failed to load state: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

def save_professor_names():
    """Save professor names to a JSON string."""
    professor_names = sorted(list(st.session_state.professor_data.index))
    return json.dumps({"professor_names": professor_names}, indent=2)

def load_professor_names(names_json):
    """Load professor names from JSON and initialize them."""
    try:
        data = json.loads(names_json)
        professor_names = sorted(data["professor_names"])
        
        # Clear existing professor data
        st.session_state.professor_data = pd.DataFrame(columns=[f"Slot {i + 1}" for i in range(st.session_state.num_slots)])
        
        # Add each professor with default availability (all available)
        for name in professor_names:
            availability = [1] * st.session_state.num_slots
            st.session_state.professor_data.loc[name] = availability
        
        st.rerun()
    except Exception as e:
        st.error(f"Failed to load professor names: {str(e)}")

def optimize_schedule():
    """Run the optimization algorithm."""
    if st.session_state.professor_data.empty:
        st.warning("Please add at least one professor first.")
        return
    
    if st.session_state.visitor_data.empty:
        st.warning("Please add at least one visitor first.")
        return
    
    # Check if any visitors have preferences
    has_preferences = False
    for visitor in st.session_state.visitor_data.index:
        visitor_prefs = st.session_state.visitor_data.loc[visitor]
        if any(pd.notna(pref) and str(pref).strip() != '' and str(pref).strip() != ' ' for pref in visitor_prefs):
            has_preferences = True
            break
    
    if not has_preferences:
        st.warning("Please add at least one visitor preference before optimizing.")
        return
    
    time_slots = [f"Slot {i+1}" for i in range(st.session_state.num_slots)]
    max_students = st.session_state.get('max_students', 1)
    
    try:
        visitor_schedule, professor_schedule, preference_analysis = sop.generate_schedule(
            st.session_state.visitor_data.copy(),
            st.session_state.professor_data.copy(),
            time_slots,
            max_students=max_students
        )
        
        st.session_state.visitor_schedule = visitor_schedule
        st.session_state.professor_schedule = professor_schedule
        st.session_state.preference_analysis = preference_analysis
        st.success("Schedule optimized successfully!")
        st.rerun()
    except Exception as e:
        st.error(f"Optimization failed: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

def export_to_excel():
    """Export schedules to Excel format with alphabetical sorting and color coding."""
    if st.session_state.visitor_schedule is None or st.session_state.professor_schedule is None:
        st.warning("Please optimize the schedule first.")
        return None
    
    output = BytesIO()
    
    # Sort schedules alphabetically
    visitor_schedule_sorted = st.session_state.visitor_schedule.sort_index()
    professor_schedule_sorted = st.session_state.professor_schedule.sort_index()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write visitor schedule
        visitor_schedule_sorted.to_excel(writer, sheet_name="Visitor Schedule")
        
        # Write professor schedule
        professor_schedule_sorted.to_excel(writer, sheet_name="Professor Schedule")
        
        # Apply color coding to professor schedule
        workbook = writer.book
        worksheet = writer.sheets["Professor Schedule"]
        
        # Import openpyxl styles
        from openpyxl.styles import PatternFill
        
        # Define colors
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Get professor data for availability checking
        # Map time slot column names (Time Slot 1 -> Slot 1)
        for row_idx, prof_name in enumerate(professor_schedule_sorted.index, start=2):  # start=2 because row 1 is header
            if prof_name in st.session_state.professor_data.index:
                for col_idx, col_name in enumerate(professor_schedule_sorted.columns, start=2):  # start=2 because col 1 is professor name
                    if col_name.startswith("Time Slot "):
                        slot_num = col_name.replace("Time Slot ", "")
                        slot_col = f"Slot {slot_num}"
                        
                        # Check if slot column exists in professor_data
                        if slot_col in st.session_state.professor_data.columns:
                            availability = st.session_state.professor_data.loc[prof_name, slot_col]
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            
                            if availability == 1:
                                cell.fill = green_fill  # Green for available
                            else:
                                cell.fill = red_fill  # Red for unavailable
    
    output.seek(0)
    return output

# Main app
st.title("Meeting Scheduler")

# Create tabs
tab1, tab2, tab3, tab4 = st.tabs(["Initialize", "Faculty Info", "Visitors Info", "Schedules"])

# Tab 1: Initialize
with tab1:
    st.header("Initialize Schedule")
    
    col1, col2 = st.columns([3, 1])
    with col1:
        st.subheader("Save/Load Progress")
        col_save, col_load = st.columns(2)
        
        with col_save:
            # Initialize filename in session state
            if 'progress_save_filename' not in st.session_state:
                st.session_state.progress_save_filename = "meeting_scheduler_state.json"
            
            # Custom filename input
            custom_filename = st.text_input(
                "Filename (without extension):",
                value=st.session_state.progress_save_filename.replace(".json", ""),
                key="progress_filename_input",
                help="Enter a custom filename for saving progress"
            )
            # Update session state
            if custom_filename:
                # Ensure it has .json extension
                filename = custom_filename if custom_filename.endswith('.json') else f"{custom_filename}.json"
                st.session_state.progress_save_filename = filename
            else:
                st.session_state.progress_save_filename = "meeting_scheduler_state.json"
            
            state_json = save_state()
            st.download_button(
                label="Save Progress",
                data=state_json,
                file_name=st.session_state.progress_save_filename,
                mime="application/json"
            )
        
        with col_load:
            uploaded_file = st.file_uploader("Load Progress", type=['json'], key="load_state")
            if uploaded_file is not None:
                state_content = uploaded_file.read().decode('utf-8')
                if st.button("Load", key="load_button"):
                    load_state(state_content)
    
    st.divider()
    
    st.subheader("Number of Time Slots")
    num_slots = st.selectbox(
        "Select number of time slots:",
        options=list(range(1, 11)),
        index=st.session_state.num_slots - 1 if st.session_state.num_slots <= 10 else 0,
        key="slot_selector"
    )
    
    if st.button("Initialize" if not st.session_state.initialized else "Update", key="init_button"):
        initialize_schedule(num_slots)
        st.success(f"Schedule initialized with {num_slots} time slots!")

# Tab 2: Faculty Info
with tab2:
    if not st.session_state.initialized:
        st.warning("Please initialize the schedule first in the 'Initialize' tab.")
    else:
        st.header("Faculty Information")
        
        # Save/Load Professor Names
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Save Professors")
            # Initialize filename in session state
            if 'prof_save_filename' not in st.session_state:
                st.session_state.prof_save_filename = "professor_names.json"
            
            # Custom filename input
            custom_filename = st.text_input(
                "Filename (without extension):",
                value=st.session_state.prof_save_filename.replace(".json", ""),
                key="prof_filename_input",
                help="Enter a custom filename for saving professors"
            )
            # Update session state
            if custom_filename:
                # Ensure it has .json extension
                filename = custom_filename if custom_filename.endswith('.json') else f"{custom_filename}.json"
                st.session_state.prof_save_filename = filename
            else:
                st.session_state.prof_save_filename = "professor_names.json"
            
            prof_names_json = save_professor_names()
            st.download_button(
                label="Save Professors",
                data=prof_names_json,
                file_name=st.session_state.prof_save_filename,
                mime="application/json"
            )
        
        with col2:
            uploaded_profs = st.file_uploader("Load Professors", type=['json'], key="load_profs")
            if uploaded_profs is not None:
                prof_content = uploaded_profs.read().decode('utf-8')
                if st.button("Load Professors", key="load_profs_button"):
                    load_professor_names(prof_content)
        
        st.divider()
        
        # Add Professor Form
        with st.expander("Add New Professor", expanded=False):
            # Use a form to handle input clearing better
            with st.form("add_professor_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                with col1:
                    prof_first_name = st.text_input("First Name", key="prof_first")
                with col2:
                    prof_last_name = st.text_input("Last Name", key="prof_last")
                
                st.write("Unavailable Time Slots:")
                unavailable_slots = []
                cols = st.columns(min(6, st.session_state.num_slots))
                for i in range(st.session_state.num_slots):
                    col_idx = i % len(cols)
                    with cols[col_idx]:
                        checkbox_key = f"prof_unavailable_{i}"
                        # Streamlit stores widget state in session_state[key], so we can read it directly
                        checked = st.checkbox(f"Slot {i+1}", key=checkbox_key)
                        if checked:
                            unavailable_slots.append(i)
                
                submitted = st.form_submit_button("Add Professor")
                
                if submitted:
                    if prof_first_name and prof_last_name:
                        add_professor(prof_first_name, prof_last_name, unavailable_slots)
                        st.success(f"Professor {prof_last_name.upper()}, {prof_first_name.upper()} added!")
                        st.rerun()
                    else:
                        st.warning("Please enter both first and last name.")
        
        st.divider()
        
        # Professor Table
        if not st.session_state.professor_data.empty:
            st.subheader("Professors")
            
            # Display professor table with availability (sorted alphabetically)
            sorted_professors = sorted(st.session_state.professor_data.index)
            for prof_idx, prof_name in enumerate(sorted_professors):
                with st.expander(f"Professor: {prof_name}", expanded=False):
                    # Allow renaming
                    new_name = st.text_input("Professor Name", value=prof_name, key=f"rename_prof_{prof_idx}")
                    if new_name != prof_name and new_name:
                        if new_name not in st.session_state.professor_data.index:
                            st.session_state.professor_data.rename(index={prof_name: new_name}, inplace=True)
                            # Update visitor preferences
                            for visitor in st.session_state.visitor_data.index:
                                for col in st.session_state.visitor_data.columns:
                                    if st.session_state.visitor_data.loc[visitor, col] == prof_name:
                                        st.session_state.visitor_data.loc[visitor, col] = new_name
                            st.rerun()
                    
                    # Availability selectors
                    cols = st.columns(min(6, st.session_state.num_slots))
                    for slot_idx in range(st.session_state.num_slots):
                        col_idx = slot_idx % len(cols)
                        with cols[col_idx]:
                            current_avail = st.session_state.professor_data.loc[prof_name, f"Slot {slot_idx + 1}"]
                            selected = st.selectbox(
                                f"Slot {slot_idx + 1}",
                                options=["Available", "Unavailable"],
                                index=0 if current_avail == 1 else 1,
                                key=f"prof_avail_{prof_idx}_{slot_idx}"
                            )
                            new_value = 1 if selected == "Available" else 0
                            if new_value != current_avail:
                                st.session_state.professor_data.loc[prof_name, f"Slot {slot_idx + 1}"] = new_value
                                # Check if professor became completely unavailable and clear preferences
                                clear_preferences_for_unavailable_professors()
                    
                    # Remove button
                    if st.button("Remove Professor", key=f"remove_prof_exp_{prof_idx}"):
                        remove_professor(prof_name)
                        st.rerun()
            
            # Also show a summary table (sorted alphabetically)
            st.subheader("Summary Table")
            prof_summary = st.session_state.professor_data.copy()
            # Sort by index (professor names) alphabetically
            prof_summary = prof_summary.sort_index()
            prof_summary_display = prof_summary.copy()
            for col in prof_summary_display.columns:
                prof_summary_display[col] = prof_summary_display[col].map({1: "Available", 0: "Unavailable"})
            
            # Apply color coding: green for Available, red for Unavailable
            def color_cells(val):
                if val == "Available":
                    return 'background-color: #90EE90'  # Light green
                elif val == "Unavailable":
                    return 'background-color: #FFB6C1'  # Light red
                return ''
            
            styled_prof_summary = prof_summary_display.style.applymap(color_cells)
            st.dataframe(styled_prof_summary, use_container_width=True)
            
        else:
            st.info("No professors added yet. Use the 'Add New Professor' form above.")

# Tab 3: Visitors Info
with tab3:
    if not st.session_state.initialized:
        st.warning("Please initialize the schedule first in the 'Initialize' tab.")
    else:
        st.header("Visitors Information")
        
        # Add Visitor
        with st.expander("Add New Visitor", expanded=False):
            # Use a form to handle input clearing better
            with st.form("add_visitor_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                with col1:
                    visitor_first_name = st.text_input("First Name", key="visitor_first_name")
                with col2:
                    visitor_last_name = st.text_input("Last Name", key="visitor_last_name")
                
                submitted = st.form_submit_button("Add Visitor")
                
                if submitted:
                    if visitor_first_name and visitor_last_name:
                        add_visitor(visitor_first_name.strip(), visitor_last_name.strip())
                        visitor_name = f"{visitor_last_name.strip().upper()}, {visitor_first_name.strip().upper()}"
                        st.success(f"Visitor {visitor_name} added!")
                        st.rerun()
                    else:
                        st.warning("Please enter both first and last name.")
        
        # Import Visitors from Excel
        with st.expander("Import Visitors from Excel", expanded=False):
            st.info("📋 **Supported formats:**\n"
                   "- Separate columns: 'First Name' and 'Last Name' (or variations)\n"
                   "- Single column: 'Name' (will be split by comma or space)\n"
                   "- First two columns will be used if no recognized headers found")
            
            uploaded_file = st.file_uploader(
                "Choose an Excel file (.xlsx or .xls)",
                type=['xlsx', 'xls'],
                key="visitor_excel_upload"
            )
            
            if uploaded_file is not None:
                if st.button("Import Visitors", key="import_visitors_button"):
                    success, message, errors = import_visitors_from_excel(uploaded_file)
                    
                    if success:
                        st.success(message)
                        if errors and len(errors) > 0:
                            with st.expander("View Errors", expanded=False):
                                for error in errors[:10]:  # Show first 10 errors
                                    st.text(error)
                                if len(errors) > 10:
                                    st.text(f"... and {len(errors) - 10} more errors")
                        st.rerun()
                    else:
                        st.error(message)
        
        # Delete Visitors
        if not st.session_state.visitor_data.empty:
            with st.expander("Delete Visitors", expanded=False):
                sorted_visitor_names = sorted(list(st.session_state.visitor_data.index))
                selected_visitors_to_delete = st.multiselect(
                    "Select visitors to delete:",
                    options=sorted_visitor_names,
                    key="visitors_to_delete_select",
                    help="Select one or more visitors to remove from the list"
                )
                
                if selected_visitors_to_delete:
                    st.warning(f"⚠️ You are about to delete {len(selected_visitors_to_delete)} visitor(s).")
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        if st.button("🗑️ Delete Selected", key="delete_visitors_button", type="primary"):
                            deleted_count = 0
                            # Batch delete all selected visitors
                            for visitor_name in selected_visitors_to_delete:
                                if visitor_name in st.session_state.visitor_data.index:
                                    st.session_state.visitor_data = st.session_state.visitor_data.drop(index=visitor_name)
                                    if visitor_name in st.session_state.name_memory:
                                        st.session_state.name_memory.remove(visitor_name)
                                    deleted_count += 1
                            st.success(f"Successfully deleted {deleted_count} visitor(s)!")
                            st.rerun()
                    with col2:
                        st.write("")  # Spacing
        
        st.divider()
        
        # Visitor Table
        if not st.session_state.visitor_data.empty:
            st.subheader("Visitors and Preferences")
            
            # Clear preferences for any professors who are completely unavailable
            clear_preferences_for_unavailable_professors()
            
            # Get professor names for dropdowns
            professor_names = sorted(list(st.session_state.professor_data.index)) if not st.session_state.professor_data.empty else []
            
            if not professor_names:
                st.warning("Please add professors first in the 'Faculty Info' tab.")
            else:
                # Show current professor count and list for reference with refresh option
                col_info, col_refresh = st.columns([4, 1])
                with col_info:
                    if professor_names:
                        prof_list_display = ", ".join(professor_names[:5])
                        if len(professor_names) > 5:
                            prof_list_display += f", ... (+{len(professor_names) - 5} more)"
                        st.info(f"💡 **{len(professor_names)} professor(s) available:** {prof_list_display}")
                with col_refresh:
                    if st.button("🔄 Refresh", help="Refresh to see newly added professors", key="refresh_visitor_editor"):
                        st.rerun()
                
                # Create editable table for visitors (sorted alphabetically)
                visitor_display = st.session_state.visitor_data.copy()
                # Sort by index (visitor names) alphabetically
                visitor_display = visitor_display.sort_index()
                visitor_display.index.name = "Visitor"
                # Normalize empty values: convert " " (space) to "" (empty string) for display
                for col in visitor_display.columns:
                    visitor_display[col] = visitor_display[col].apply(
                        lambda x: "" if (isinstance(x, str) and x.strip() == " ") or x == " " else x
                    )
                
                # Create column config for preferences - this gets the latest professor list on each render
                # The key is to ensure professor_names is fetched fresh each time this code runs
                column_config = {}
                current_prof_names = sorted(list(st.session_state.professor_data.index)) if not st.session_state.professor_data.empty else []
                # Filter out any empty strings or whitespace-only names
                current_prof_names = [name for name in current_prof_names if name and str(name).strip()]
                
                # Check which professors are completely unavailable and mark them
                unavailable_professors = set()
                for prof_name in current_prof_names:
                    if prof_name in st.session_state.professor_data.index:
                        # Check if all slots are unavailable (all values are 0)
                        prof_row = st.session_state.professor_data.loc[prof_name]
                        if all(prof_row[col] == 0 for col in prof_row.index):
                            unavailable_professors.add(prof_name)
                
                # Create display names with strike-through for unavailable professors
                # Use Unicode strikethrough combining characters
                prof_display_names = []
                prof_name_mapping = {}  # Map display name to original name
                for prof_name in current_prof_names:
                    if prof_name in unavailable_professors:
                        # Add strikethrough using Unicode combining characters
                        display_name = '\u0336'.join(prof_name) + '\u0336'  # Strikethrough each character
                        prof_display_names.append(display_name)
                        prof_name_mapping[display_name] = prof_name
                    else:
                        prof_display_names.append(prof_name)
                        prof_name_mapping[prof_name] = prof_name
                
                for col in visitor_display.columns:
                    # Always use the current professor_names list (updated on each page render)
                    # Use empty string instead of space for the blank option to avoid duplicates
                    column_config[col] = st.column_config.SelectboxColumn(
                        col,
                        options=[""] + prof_display_names,  # Use display names with strikethrough
                        required=False
                    )
                
                # Store the mapping in session state so we can use it when updating preferences
                st.session_state.prof_name_mapping = prof_name_mapping
                
                # Use a stable key but the column_config will have fresh options
                edited_visitors = st.data_editor(
                    visitor_display,
                    use_container_width=True,
                    key="visitor_editor",
                    column_config=column_config,
                    num_rows="dynamic"
                )
                
                # Update visitor_data based on edits
                if not edited_visitors.equals(visitor_display):
                    # Handle new rows
                    for visitor in edited_visitors.index:
                        if visitor not in st.session_state.visitor_data.index:
                            # New visitor added
                            if visitor not in st.session_state.name_memory:
                                st.session_state.name_memory.append(visitor)
                        # Update preferences
                        for col in edited_visitors.columns:
                            if col in st.session_state.visitor_data.columns:
                                pref_value = edited_visitors.loc[visitor, col]
                                # Convert empty string to space for consistency with optimization
                                if pref_value == "" or (isinstance(pref_value, str) and pref_value.strip() == ""):
                                    pref_value = " "
                                else:
                                    # If the value is a display name with strikethrough, convert back to original name
                                    prof_mapping = st.session_state.get('prof_name_mapping', {})
                                    if pref_value in prof_mapping:
                                        pref_value = prof_mapping[pref_value]
                                st.session_state.visitor_data.loc[visitor, col] = pref_value
                    
                    # Handle removed rows
                    for visitor in st.session_state.visitor_data.index:
                        if visitor not in edited_visitors.index:
                            remove_visitor(visitor)
                            break
                    
                    st.rerun()
        else:
            st.info("No visitors added yet. Use the 'Add New Visitor' form above.")

# Tab 4: Schedules
with tab4:
    if not st.session_state.initialized:
        st.warning("Please initialize the schedule first in the 'Initialize' tab.")
    else:
        st.header("Optimized Schedules")
        
        # Optimization controls
        col1, col2 = st.columns([2, 1])
        with col1:
            max_students = st.number_input(
                "Students per Professor per Slot:",
                min_value=1,
                max_value=5,
                value=st.session_state.get('max_students', 1),
                key="max_students_input"
            )
            st.session_state.max_students = max_students
        
        with col2:
            st.write("")  # Spacing
            st.write("")  # Spacing
            if st.button("Optimize Schedule", type="primary", key="optimize_button"):
                optimize_schedule()
        
        st.divider()
        
        # Display schedules
        if st.session_state.visitor_schedule is not None and st.session_state.professor_schedule is not None:
            st.subheader("Visitor Schedule")
            # Sort visitor schedule alphabetically
            visitor_schedule_display = st.session_state.visitor_schedule.copy()
            visitor_schedule_display = visitor_schedule_display.sort_index()
            st.dataframe(visitor_schedule_display, use_container_width=True)
            
            st.divider()
            
            st.subheader("Professor Schedule")
            # Color code based on professor availability
            # Sort professor schedule alphabetically
            prof_schedule_display = st.session_state.professor_schedule.copy()
            prof_schedule_display = prof_schedule_display.sort_index()
            
            def color_prof_schedule_cells(series):
                """Color cells based on professor availability for that time slot."""
                prof_name = series.name
                styles = []
                
                for col_name in prof_schedule_display.columns:
                    # Map column name to slot number (e.g., "Time Slot 1" -> "Slot 1")
                    if col_name.startswith("Time Slot "):
                        slot_num = col_name.replace("Time Slot ", "")
                        slot_col = f"Slot {slot_num}"
                        
                        # Check professor availability
                        if prof_name in st.session_state.professor_data.index:
                            # Check if the slot column exists in professor_data
                            if slot_col in st.session_state.professor_data.columns:
                                availability = st.session_state.professor_data.loc[prof_name, slot_col]
                                if availability == 1:
                                    styles.append('background-color: #90EE90')  # Light green for available
                                else:
                                    styles.append('background-color: #FFB6C1')  # Light red for unavailable
                            else:
                                # Slot doesn't exist in current professor_data (maybe schedule was generated with different num_slots)
                                styles.append('')  # No styling if slot column doesn't exist
                        else:
                            styles.append('')  # No styling if professor not found
                    else:
                        styles.append('')  # No styling for non-time-slot columns
                
                return styles
            
            # Apply styling row by row
            styled_prof_schedule = prof_schedule_display.style.apply(color_prof_schedule_cells, axis=1)
            st.dataframe(styled_prof_schedule, use_container_width=True)
            
            # Color-coded visitor preferences table
            st.subheader("Visitor Preferences Status")
            if not st.session_state.visitor_data.empty:
                # Create a styled view showing which preferences are met
                visitor_prefs_display = st.session_state.visitor_data.copy()
                visitor_prefs_display = visitor_prefs_display.sort_index()
                # Normalize empty values for display
                for col in visitor_prefs_display.columns:
                    visitor_prefs_display[col] = visitor_prefs_display[col].apply(
                        lambda x: "" if (isinstance(x, str) and x.strip() == " ") or x == " " else x
                    )
                
                def color_visitor_prefs(row):
                    """Color code preferences based on whether they're met in the schedule."""
                    visitor_name = row.name
                    styles = []
                    
                    # Get all professors this visitor is scheduled with
                    if visitor_name in st.session_state.visitor_schedule.index:
                        scheduled_profs = set()
                        for col in st.session_state.visitor_schedule.columns:
                            prof = st.session_state.visitor_schedule.loc[visitor_name, col]
                            if pd.notna(prof) and str(prof).strip() != '':
                                scheduled_profs.add(str(prof).strip())
                    else:
                        scheduled_profs = set()
                    
                    # Check each preference column
                    for col in visitor_prefs_display.columns:
                        if col.startswith('Preference'):
                            pref_prof = row[col]
                            if pref_prof and str(pref_prof).strip() != '':
                                pref_prof = str(pref_prof).strip()
                                if pref_prof in scheduled_profs:
                                    styles.append('background-color: #90EE90')  # Green for met
                                else:
                                    styles.append('background-color: #FFB6C1')  # Red for not met
                            else:
                                styles.append('')  # No color for empty preferences
                        else:
                            styles.append('')  # No color for non-preference columns
                    
                    return styles
                
                # Apply styling
                styled_visitor_prefs = visitor_prefs_display.style.apply(color_visitor_prefs, axis=1)
                st.dataframe(styled_visitor_prefs, use_container_width=True, height=300)
            
            st.divider()
            
            # Preference analysis (sorted alphabetically)
            if 'preference_analysis' in st.session_state:
                st.subheader("Preference Fulfillment Analysis")
                preference_analysis_display = st.session_state.preference_analysis.copy()
                preference_analysis_display = preference_analysis_display.sort_index()
                st.dataframe(preference_analysis_display, use_container_width=True)
            
            # Export buttons
            st.divider()
            st.subheader("Export Schedules")
            
            # Initialize filename in session state
            if 'excel_export_filename' not in st.session_state:
                st.session_state.excel_export_filename = "meeting_schedules.xlsx"
            
            # Custom filename input
            custom_excel_filename = st.text_input(
                "Filename (without extension):",
                value=st.session_state.excel_export_filename.replace(".xlsx", ""),
                key="excel_filename_input",
                help="Enter a custom filename for exporting schedules"
            )
            # Update session state
            if custom_excel_filename:
                # Ensure it has .xlsx extension
                filename = custom_excel_filename if custom_excel_filename.endswith('.xlsx') else f"{custom_excel_filename}.xlsx"
                st.session_state.excel_export_filename = filename
            else:
                st.session_state.excel_export_filename = "meeting_schedules.xlsx"
            
            col1, col2 = st.columns(2)
            with col1:
                excel_file = export_to_excel()
                if excel_file:
                    st.download_button(
                        label="Export Schedules to Excel",
                        data=excel_file,
                        file_name=st.session_state.excel_export_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_excel"
                    )
        else:
            st.info("Click 'Optimize Schedule' to generate the meeting schedules.")

